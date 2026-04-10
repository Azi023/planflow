#!/usr/bin/env python3
"""
PlanFlow Data Import — Phase 1: Excel Scanner & Normalizer

Reads all Excel files from data-import/raw/, classifies them, extracts
structured data, normalizes it, and writes clean JSON to data-import/output/.

READ-ONLY: Never modifies source files.
"""

import json
import logging
import os
import re
import sys
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).parent
RAW_DIR = BASE_DIR / "raw"
OUT_DIR = BASE_DIR / "output"
LOG_FILE = OUT_DIR / "scanner.log"

# ---------------------------------------------------------------------------
# Column alias map  (canonical_name → [list of known header variants])
# ---------------------------------------------------------------------------
COLUMN_ALIASES: dict[str, list[str]] = {
    "campaign_name": [
        "campaign name", "campaign", "campaign name ", "campain name",
    ],
    "platform": [
        "platform", "platform ", "channel", "channels", "platform  ",
    ],
    "audience_name": [
        "audience", "audience ", "audiences", "targeting audience",
        "target audience", "audience name", "audiences ",
    ],
    "audience_size": [
        "estimated audience size", "audience size", "audience size ",
        "estimated audience size ", "est. audience size", "est audience size",
        "audience size(estimated)", "est audience size ",
    ],
    "targeting": [
        "detailed targeting", "targeting audience", "targeting criteria",
        "target audience", "targeting criteria ", "detaild targeting",
        "targeting criteria  ", "detailed targeting ",
    ],
    "creative": [
        "ad type & no of content pieces", "ad type &no of content pieces ",
        "ad type &no of content pieces", "creative", "assets", "content",
        "no.of content", "no. of content", "ad type", "ad type ",
        "ad type & no.of content pieces", "ad type&no of content pieces",
        "ad type &no of content  pieces",
    ],
    "budget": [
        "budget (lkr)", "budget", "budget (lkr) ", "budget(lkr)",
        "investment", "investment (lkr)", "amount", "budget (usd)",
        "budget  (lkr)", "budget(lkr) ", "spend amount ($)",
        "amount spend (lkr)", "amount spend (lkr) ",
        "budget allocation", "budget allocation ", "budget allocation (lkr)",
        "allocated budget", "allocated budget (lkr)",
    ],
    "total_budget": [
        "total budget (lkr)", "total budget with management fee (lkr)",
        "total budget with mgt & asp fee", "total cost (lkr)",
        "total budget", "total cost", "total budget (usd)",
        "total budget with management", "total budget with manage",
        "total budget(lkr)", "total budget with mgt fee",
        "total budget with management fee", "total budget ",
    ],
    "media_spend": [
        "media spend (lkr)", "media spend (lkr) ", "media spend",
        "media spend ( lkr )", "amount spend (lkr)", "net media spend",
        "media spend(lkr)", "media spend  (lkr)", "media spend (usd)",
    ],
    "percentage": [
        "percentage (%)", "precentage (%)", "percentage", "%", "allocation %",
        "budget %", "split %", "precentage  (%)", "percentage  (%)",
        "percentage(%)", "precentage(%)",
    ],
    "campaign_period": [
        "campaign period", "campaign period ", "campaign period (days/week)",
        "campaign period (days)", "duration", "period",
        "campaign period (days/weeks)", "campaign period (days/ week)",
        "campaign period (days/week) ",
    ],
    "reach": [
        "reach", "reach ", "estimated reach", "estimated reach ",
        "reach pki", "reach kpi", "reach kpi ", "estimated reach kpi",
        "reach (estimated)", "reach pkI", "recch", "recch ",
    ],
    "impressions": [
        "impressions", "impressions ", "estimated impressions",
        "impressions kpi", "est. impressions", "impressions kpi ",
        "estimated impressions ", "impressions kpi  ",
    ],
    "frequency": ["frequency", "frequency ", "freq", "avg frequency", "frequency  "],
    "cpm": [
        "cpm", "cpm ", "cpm (lkr)", "cost per 1000 impressions",
        "est.cpm", "est. cpm", "avg cost (lkr)", "avg cost",
        "estimated cpm", "cpm(lkr)", "cpM (lkr)",
    ],
    "cpc": ["cpc", "cpc ", "cpc (lkr)", "cost per click", "est.cpc", "est. cpc"],
    "cpl": ["cpl", "cpl ", "cpl (lkr)", "cost per lead", "est.cpl", "est. cpl"],
    "cpr": [
        "cpr", "cpr (lkr)", "cost per reach", "est.cpr",
        "range given by platforms cpl", "cpr (lkr) ",
    ],
    "cpv": [
        "cpv", "cpv ", "cpv (lkr)", "cost per view", "cpv (2sec)", "cpv (trueview)",
        "cpv (2 sec)", "cpv(lkr)",
    ],
    "ctr": [
        "ctr", "ctr ", "ctr%", "click through rate", "click-through rate",
        "ctr (estimated)", "ctr%  ",
    ],
    "video_views": [
        "video views", "video views ", "views", "3-second video views",
        "thruplay", "thruplays", "views kpi", "views kpi ",
        "3-sec video views", "video view kpi",
    ],
    "leads": ["leads", "leads ", "results", "conversions", "no of leads", "leads kpi"],
    "clicks": [
        "clicks", "clicks ", "link clicks", "all clicks", "website clicks",
        "clicks kpi", "link clicks kpi",
    ],
    "country": ["country", "country ", "location", "geo", "region", "country  "],
    "buy_type": [
        "buy type", "buy type ", "buying type", "auction/cpm", "buying model",
        "buy type  ",
    ],
    "reference_number": [
        "reference no", "ref no", "ref #", "reference number",
        "ref.no", "reference no.", "ref no.", "reference no:",
    ],
    "prepared_by": ["prepared by", "prepared by ", "created by", "planner"],
    "mgmt_fee": [
        "management fee", "mgmt fee", "mgt fee", "agency fee",
        "management fee ", "agency commission", "mgmt fee ", "mgt fee ",
    ],
    "range_cpm": [
        "range given by platforms cpm", "range cpm", "platform range cpm",
        "cpm range", "range given by platforms",
    ],
    "range_cpl": [
        "range given by platforms cpl", "range cpl", "platform range cpl",
        "cpl range",
    ],
    "landing_page_views": [
        "landing page views", "lpv", "landing page view", "lp views",
        "landing page views kpi", "lpv kpi",
    ],
    "estimated_deliveries": [
        "estimated deliveries", "estimated deliveries ", "est. deliveries",
        "estimated delivery", "deliveries", "est deliveries",
    ],
    "avg_cost": [
        "avg cost", "avg cost ", "avg. cost", "average cost",
        "avg cost (lkr)", "unit cost", "cost per unit",
    ],
    "engagements": [
        "engagements", "engagement", "post engagements", "total engagements",
        "engagement kpi", "engagements kpi", "engagements kpi ",
        "engagement kpi ", "post engagement kpi",
    ],
    "page_likes": [
        "page likes", "page like", "per page like", "page like cost",
        "page likes kpi",
    ],
    "objective": [
        "channels / objectives", "channels / objectives ", "channels/objectives",
        "objective", "objective ", "channels / objective", "objective  ",
    ],
    "date": ["date", "date ", "campaign date", "plan date"],
    "client": ["client", "client ", "client name", "advertiser"],
    "prepared_date": ["date", "date "],
}

# Build a quick lookup: normalized_alias → canonical
_ALIAS_LOOKUP: dict[str, str] = {}
for canonical, aliases in COLUMN_ALIASES.items():
    for alias in aliases:
        key = alias.strip().lower()
        if key not in _ALIAS_LOOKUP:
            _ALIAS_LOOKUP[key] = canonical

# ---------------------------------------------------------------------------
# Platform normalization map
# ---------------------------------------------------------------------------
PLATFORM_NORMALIZE: dict[str, str] = {
    # Meta family
    "fb/ig": "meta_ig",
    "fb & ig": "meta_ig",
    "fb+ig": "meta_ig",
    "facebook/instagram": "meta_ig",
    "meta(facebook/instagram)": "meta_ig",
    "meta(facebok/instagram)": "meta_ig",
    "meta + ig": "meta_ig",
    "meta+ig": "meta_ig",
    "facebook + instagram": "meta_ig",
    "facebook/instagram": "meta_ig",
    "fb+ig": "meta_ig",
    "facebook & instagram": "meta_ig",
    "meta & ig": "meta_ig",
    "meta (facebook/instagram)": "meta_ig",
    "meta(facebook / instagram)": "meta_ig",
    "facebook": "meta",
    "meta only": "meta",
    "meta": "meta",
    "instagram": "ig",
    "ig only": "ig",
    "ig": "ig",
    "ig follower": "ig_follower",
    "instagram follower": "ig_follower",
    "instagram followers": "ig_follower",
    "page like": "meta_page_like",
    "meta page like": "meta_page_like",
    "page likes": "meta_page_like",
    "fb page like": "meta_page_like",
    "facebook page like": "meta_page_like",
    # Google family
    "gdn": "gdn",
    "google display": "gdn",
    "google display network": "gdn",
    "google (gdn)": "gdn",
    "google (gdn) rmkt": "gdn",
    "google gdn": "gdn",
    "display": "gdn",
    "youtube": "youtube_video",
    "youtube video views": "youtube_video",
    "youtube video": "youtube_video",
    "yt video views": "youtube_video",
    "youtube (video views)": "youtube_video",
    "google (yt)": "youtube_video",
    "google yt": "youtube_video",
    "google youtube": "youtube_video",
    "youtube bumper": "youtube_bumper",
    "yt bumper": "youtube_bumper",
    "bumper": "youtube_bumper",
    "bumper ads": "youtube_bumper",
    "search": "search",
    "google search": "search",
    "google ads search": "search",
    "google ads - search": "search",
    "google (search)": "search",
    "demand gen": "demand_gen",
    "demand generation": "demand_gen",
    "google demand gen": "demand_gen",
    "performance max": "perf_max",
    "pmax": "perf_max",
    "perf max": "perf_max",
    "performance max (pmax)": "perf_max",
    "google performance max": "perf_max",
    # TikTok
    "tiktok": "tiktok",
    "tik tok": "tiktok",
    "tik-tok": "tiktok",
    # LinkedIn
    "linkedin": "linkedin",
    "linked in": "linkedin",
    "linked-in": "linkedin",
}

# Known PlanFlow platforms — anything not here is "new"
KNOWN_PLATFORMS = {
    "meta_ig", "meta", "ig", "ig_follower", "meta_page_like",
    "gdn", "youtube_video", "youtube_bumper", "search", "demand_gen", "perf_max",
    "tiktok", "linkedin",
}

# ---------------------------------------------------------------------------
# Logging setup
# ---------------------------------------------------------------------------
def setup_logging() -> logging.Logger:
    logger = logging.getLogger("scanner")
    logger.setLevel(logging.DEBUG)
    fmt = logging.Formatter("[%(levelname)s] %(message)s")

    console = logging.StreamHandler(sys.stdout)
    console.setLevel(logging.INFO)
    console.setFormatter(fmt)

    fh = logging.FileHandler(LOG_FILE, mode="w", encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)

    logger.addHandler(console)
    logger.addHandler(fh)
    return logger


log: logging.Logger = None  # set in main()

# ---------------------------------------------------------------------------
# Column normalization
# ---------------------------------------------------------------------------
def normalize_column(raw: str) -> Optional[str]:
    """Map a raw column header to its canonical name."""
    if not raw:
        return None
    key = str(raw).strip().lower()
    return _ALIAS_LOOKUP.get(key)


# Values that look like objectives/KPIs — NOT platforms
_OBJECTIVE_VALUES = {
    "reach", "awareness", "engagement", "engagements", "traffic", "leads",
    "video views", "views", "clicks", "impressions", "cpm", "cpc", "cpl",
    "cpv", "ctr", "conversions", "thruplays", "thruplay",
}

# Values that are clearly column header labels or noise — NOT platforms
_REJECTED_PLATFORM_VALUES = {
    "platform", "creative", "creatives", "channel", "channels", "header",
    "media", "digital", "n/a", "na", "–", "—", "-",
}


def normalize_platform(raw: str) -> tuple[str, bool]:
    """
    Returns (normalized_platform, is_new).
    is_new=True when the platform isn't in KNOWN_PLATFORMS.
    Returns (None, False) when raw looks like an objective value, not a platform.
    """
    if not raw:
        return raw, False
    key = str(raw).strip().lower()
    key = re.sub(r"\s+", " ", key)

    # Reject values that are clearly objective/KPI labels or column headers, not platforms
    if key in _OBJECTIVE_VALUES or key in _REJECTED_PLATFORM_VALUES:
        return None, False

    # Reject very long strings — footnotes/disclaimers mistakenly in platform column
    if len(key) > 80:
        return None, False

    # Reject strings that look like content counts (e.g. "7 published contents")
    if re.match(r"^\d+\s+published", key) or re.match(r"^\d+\s+content", key):
        return None, False

    normalized = PLATFORM_NORMALIZE.get(key)
    if normalized is None:
        # Try partial match for compound strings like "Meta(Facebook/Instagram)"
        for k, v in PLATFORM_NORMALIZE.items():
            if len(k) > 3 and (k in key or key in k):
                normalized = v
                break
    if normalized is None:
        normalized = key.replace(" ", "_")
    is_new = normalized not in KNOWN_PLATFORMS
    return normalized, is_new


# ---------------------------------------------------------------------------
# Number / range parsing
# ---------------------------------------------------------------------------
def parse_number(s: Any) -> Optional[float]:
    """Parse a scalar value to float, handling K/M suffixes, currency symbols, commas."""
    if s is None:
        return None
    raw = str(s).strip()
    # Remove known currency prefixes / noise
    raw = re.sub(r"(?i)(lkr|usd|\$|rs\.?|rs\s)", "", raw)
    raw = raw.replace(",", "").strip()
    # Remove trailing/leading dash placeholders
    if raw in ("–", "—", "-", "N/A", "n/a", "", "0.0", "−"):
        return None

    # Suffix multiplier
    multiplier = 1.0
    upper = raw.upper()
    if upper.endswith("MN"):
        multiplier = 1_000_000
        raw = raw[:-2].strip()
    elif upper.endswith("M"):
        multiplier = 1_000_000
        raw = raw[:-1].strip()
    elif upper.endswith("K"):
        multiplier = 1_000
        raw = raw[:-1].strip()

    # Parenthetical explanations like "320K-420K (ThruPlay)" — strip them
    raw = re.sub(r"\(.*\)", "", raw).strip()
    raw = re.sub(r"\[.*\]", "", raw).strip()

    if not raw:
        return None

    try:
        return float(raw) * multiplier
    except ValueError:
        return None


def parse_range(val: Any) -> tuple[Optional[float], Optional[float]]:
    """
    Parse a range string into (low, high).
    Handles: "35 - 65", "5.6M-5.5M", "2.5M - 3M", "330", "2 times per week"
    """
    if val is None:
        return None, None
    s = str(val).strip()

    # Strip leading/trailing noise
    s = re.sub(r"(?i)(lkr|usd|\$)", "", s).strip()

    if not s or s in ("–", "—", "-", "N/A", "n/a", ""):
        return None, None

    # Try splitting on range separators — order matters (longer first)
    for sep in [" - ", " – ", " — ", "–", "—"]:
        if sep in s:
            parts = s.split(sep, 1)
            a = parse_number(parts[0].strip())
            b = parse_number(parts[1].strip())
            if a is not None and b is not None:
                if a > b:
                    log.debug(f"  Auto-fixed inverted range '{val}' → ({b}, {a})")
                return (min(a, b), max(a, b)) if (a is not None and b is not None) else (a, b)
            return a, b

    # Plain hyphen as separator (must not be negative number or K-suffix)
    # Only try if there are digits on both sides of the hyphen
    m = re.match(r"^([\d.,]+[KkMm]?)\s*-\s*([\d.,]+[KkMm]?.*)$", s)
    if m:
        a = parse_number(m.group(1))
        b = parse_number(m.group(2))
        if a is not None and b is not None:
            if a > b:
                log.debug(f"  Auto-fixed inverted range '{val}' → ({b}, {a})")
            return (min(a, b), max(a, b))
        return a, b

    # Single value
    n = parse_number(s)
    return n, n


def parse_currency(val: Any) -> str:
    """Detect currency from a cell value."""
    if val is None:
        return "LKR"
    s = str(val).upper()
    if "USD" in s or "$" in s:
        return "USD"
    return "LKR"


def parse_pct(val: Any) -> Optional[float]:
    """Parse a percentage value (e.g. 0.15 or '15%') to a decimal like 15.0."""
    if val is None:
        return None
    s = str(val).strip().replace("%", "").strip()
    try:
        n = float(s)
        # Distinguish between 0.15 (fraction) and 15 (already %)
        if 0 < n < 1:
            return round(n * 100, 4)
        return round(n, 4)
    except ValueError:
        return None


def parse_number_aggressive(s: str) -> Optional[float]:
    """Parse numbers with K/M/B suffixes, commas, and various formats."""
    if not s:
        return None
    s = str(s).strip().replace(",", "").replace(" ", "")

    multiplier = 1
    s_upper = s.upper()
    if s_upper.endswith("B"):
        multiplier = 1_000_000_000
        s = s[:-1]
    elif s_upper.endswith("MN"):
        multiplier = 1_000_000
        s = s[:-2]
    elif s_upper.endswith("M"):
        multiplier = 1_000_000
        s = s[:-1]
    elif s_upper.endswith("K"):
        multiplier = 1_000
        s = s[:-1]

    try:
        return float(s) * multiplier
    except ValueError:
        return None


def extract_kpis_from_text(text: str) -> dict[str, Optional[float]]:
    """Extract KPI values embedded in descriptive text cells like '4.6 Mn Reach' or '320K-420K (ThruPlay)'."""
    if not text or not isinstance(text, str):
        return {}

    kpis: dict[str, Optional[float]] = {}
    text_lower = text.lower()

    def _parse_groups(groups: tuple) -> tuple[Optional[float], Optional[float]]:
        if len(groups) == 2:
            lo = parse_number_aggressive(groups[0])
            hi = parse_number_aggressive(groups[1])
            return lo, hi
        val = parse_number_aggressive(groups[0])
        return val, val

    # Reach
    for pat in [
        r"([\d,.]+[kmb]?(?:mn)?)\s*[-–]\s*([\d,.]+[kmb]?(?:mn)?)\s*reach",
        r"reach\s*[:=]?\s*([\d,.]+[kmb]?(?:mn)?)",
        r"([\d,.]+\s*(?:mn|m|k|b)?)\s*reach",
    ]:
        m = re.search(pat, text_lower)
        if m:
            lo, hi = _parse_groups(m.groups())
            if lo is not None:
                kpis["reach_low"] = lo
                kpis["reach_high"] = hi
            break

    # Impressions
    for pat in [
        r"([\d,.]+[kmb]?(?:mn)?)\s*[-–]\s*([\d,.]+[kmb]?(?:mn)?)\s*(?:impression|impr)",
        r"([\d,.]+[kmb]?(?:mn)?)\s*(?:impression|impr)",
    ]:
        m = re.search(pat, text_lower)
        if m:
            lo, hi = _parse_groups(m.groups())
            if lo is not None:
                kpis["impressions_low"] = lo
                kpis["impressions_high"] = hi
            break

    # Video views / ThruPlay
    for pat in [
        r"([\d,.]+[kmb]?(?:mn)?)\s*[-–]\s*([\d,.]+[kmb]?(?:mn)?)\s*(?:thruplay|video view|view)",
        r"([\d,.]+[kmb]?(?:mn)?)\s*(?:thruplay|video view|view)",
    ]:
        m = re.search(pat, text_lower)
        if m:
            lo, hi = _parse_groups(m.groups())
            if lo is not None:
                kpis["video_views_low"] = lo
                kpis["video_views_high"] = hi
            break

    # Engagements
    for pat in [
        r"([\d,.]+[kmb]?(?:mn)?)\s*[-–]\s*([\d,.]+[kmb]?(?:mn)?)\s*engagement",
        r"([\d,.]+[kmb]?(?:mn)?)\s*engagement",
    ]:
        m = re.search(pat, text_lower)
        if m:
            lo, hi = _parse_groups(m.groups())
            if lo is not None:
                kpis["engagements_low"] = lo
                kpis["engagements_high"] = hi
            break

    # Frequency
    freq_m = re.search(r"(?:avg\s*)?frequency\s*[:=]?\s*([\d.]+)", text_lower)
    if freq_m:
        val = float(freq_m.group(1))
        kpis["frequency_low"] = val
        kpis["frequency_high"] = val

    # Leads
    for pat in [
        r"([\d,.]+[kmb]?(?:mn)?)\s*[-–]\s*([\d,.]+[kmb]?(?:mn)?)\s*lead",
        r"([\d,.]+[kmb]?(?:mn)?)\s*lead",
    ]:
        m = re.search(pat, text_lower)
        if m:
            lo, hi = _parse_groups(m.groups())
            if lo is not None:
                kpis["leads_low"] = lo
                kpis["leads_high"] = hi
            break

    return kpis


# ---------------------------------------------------------------------------
# Header detection
# ---------------------------------------------------------------------------
def find_header_row(ws, max_scan: int = 25) -> tuple[int, dict[int, str]]:
    """
    Scan rows 1–max_scan to find the row with the most recognised columns.
    Returns (row_idx_1based, {col_idx: canonical_name}).
    """
    best_row = -1
    best_score = 0
    best_mapping: dict[int, str] = {}

    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        mapping: dict[int, str] = {}
        for c_idx, cell in enumerate(row, start=1):
            if cell is None:
                continue
            canonical = normalize_column(str(cell))
            if canonical:
                mapping[c_idx] = canonical
        score = len(mapping)
        if score > best_score:
            best_score = score
            best_row = r_idx
            best_mapping = mapping

    return best_row, best_mapping


# ---------------------------------------------------------------------------
# Metadata extraction (cells above header)
# ---------------------------------------------------------------------------
def extract_metadata(ws, header_row: int) -> dict[str, Any]:
    """Look for key:value pairs above the header row."""
    meta: dict[str, Any] = {}
    meta_keys = {
        "client": ["client", "client "],
        "prepared_by": ["prepared by", "prepared by "],
        "reference_number": [
            "reference no", "ref no", "ref.no", "ref #", "reference no:", "reference no."
        ],
        "date": ["date", "date "],
        "campaign_name": ["campaign name", "campaign name "],
    }
    # Build reverse map
    reverse_map: dict[str, str] = {}
    for canonical, variants in meta_keys.items():
        for v in variants:
            reverse_map[v.strip().lower()] = canonical

    for r_idx in range(1, min(header_row, 15)):
        row = list(ws.iter_rows(min_row=r_idx, max_row=r_idx, values_only=True))[0]
        for c_idx, cell in enumerate(row):
            if cell is None:
                continue
            key_str = str(cell).strip().lower()
            canonical = reverse_map.get(key_str)
            if canonical and canonical not in meta:
                # Look at the next cell(s) for the value
                for offset in range(1, 4):
                    if c_idx + offset < len(row):
                        val = row[c_idx + offset]
                        if val is not None and str(val).strip():
                            meta[canonical] = str(val).strip()
                            break

    return meta


# ---------------------------------------------------------------------------
# File classification
# ---------------------------------------------------------------------------
def classify_file(ws) -> str:
    """Classify worksheet as 'media_plan', 'kpi_performance', or 'unknown'."""
    headers_found: set[str] = set()
    text_values: list[str] = []

    for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
        for cell in row:
            if cell is None:
                continue
            if isinstance(cell, str):
                normalized = normalize_column(cell.strip())
                if normalized:
                    headers_found.add(normalized)
                text_values.append(cell.strip().lower())

    has_budget = bool({"budget", "media_spend", "total_budget"} & headers_found)
    # "objective" column (Channels / Objectives) serves as platform equivalent in older templates
    has_platform = bool({"platform", "objective"} & headers_found)
    has_kpis = bool({"reach", "impressions", "cpm", "clicks", "video_views"} & headers_found)
    has_actuals_text = any(
        kw in tv for tv in text_values
        for kw in ("actual", "result", "delivered", "performance", "kpi report")
    )
    is_kpi_sheet = "kpi" in " ".join(text_values[:30])

    if has_budget and has_platform:
        return "media_plan"
    if has_kpis and (has_actuals_text or is_kpi_sheet):
        return "kpi_performance"
    if has_kpis and not has_budget:
        return "kpi_performance"
    if has_budget:
        return "media_plan"
    return "unknown"


# ---------------------------------------------------------------------------
# Data row extraction
# ---------------------------------------------------------------------------
_TOTAL_LABELS = {
    "total", "sub total", "subtotal", "grand total", "total media spend",
    "total budget", "total cost", "total july", "total aug", "total july/aug",
}

# Common footnote prefixes in the first cell
_FOOTNOTE_PREFIXES = (
    "results and costs", "reach and impressions", "note:", "disclaimer:",
    "all figures", "please note", "* results", "the above", "approved",
)


def is_total_row(row: tuple) -> bool:
    """Detect TOTAL / summary / footnote rows that should not be treated as data."""
    # Only examine the first non-None cell for total/footnote labels
    for idx, cell in enumerate(row):
        if cell is None:
            continue
        s = str(cell).strip()
        sl = s.lower()

        # Known total labels in ANY column
        if sl in _TOTAL_LABELS:
            return True

        # Footnote patterns only in the first few columns
        if idx < 3:
            if any(sl.startswith(prefix) for prefix in _FOOTNOTE_PREFIXES):
                return True
            # Very long string in first column with no budget/number content
            if idx == 0 and len(s) > 100 and not any(
                p in sl for p in ("meta", "youtube", "google", "tiktok", "search", "lkr", "usd")
            ):
                return True
        break  # Only check first non-None cell for the long-string heuristic

    return False


def row_is_empty(row: tuple) -> bool:
    return all(c is None or str(c).strip() == "" for c in row)


def extract_rows(ws, header_row: int, col_map: dict[int, str]) -> list[dict[str, Any]]:
    """
    Extract data rows below the header, applying the column map.
    Handles merged/blank platform cells by carrying forward the last platform.
    """
    rows: list[dict[str, Any]] = []
    last_platform: Optional[str] = None
    last_campaign: Optional[str] = None
    last_period: Optional[str] = None
    last_country: Optional[str] = None
    last_buy_type: Optional[str] = None

    consecutive_empty = 0

    for r_idx in range(header_row + 1, ws.max_row + 1):
        row = tuple(
            cell.value for cell in ws[r_idx]
        )

        if row_is_empty(row):
            consecutive_empty += 1
            if consecutive_empty >= 3:
                break
            continue
        consecutive_empty = 0

        if is_total_row(row):
            continue

        # Map cells by canonical column name — first col wins on duplicates
        record: dict[str, Any] = {}
        for c_idx, canonical in sorted(col_map.items()):
            if c_idx <= len(row):
                val = row[c_idx - 1]
                if canonical not in record or record[canonical] is None:
                    record[canonical] = val

        # Skip rows where nothing useful was extracted
        if not any(v is not None for v in record.values()):
            continue

        # Carry-forward logic for spanned/merged cells
        if record.get("platform") is None and last_platform:
            record["platform"] = last_platform
        elif record.get("platform") is not None:
            last_platform = record["platform"]

        if record.get("campaign_name") is None and last_campaign:
            record["campaign_name"] = last_campaign
        elif record.get("campaign_name") is not None:
            last_campaign = record["campaign_name"]

        if record.get("campaign_period") is None and last_period:
            record["campaign_period"] = last_period
        elif record.get("campaign_period") is not None:
            last_period = record["campaign_period"]

        if record.get("country") is None and last_country:
            record["country"] = last_country
        elif record.get("country") is not None:
            last_country = record["country"]

        if record.get("buy_type") is None and last_buy_type:
            record["buy_type"] = last_buy_type
        elif record.get("buy_type") is not None:
            last_buy_type = record["buy_type"]

        # Skip rows that still have no platform and no budget — likely footnotes
        if record.get("platform") is None and record.get("budget") is None:
            continue

        rows.append(record)

    return rows


# ---------------------------------------------------------------------------
# Platform inference from "Channels/Objectives" combined column values
# ---------------------------------------------------------------------------
def _infer_platform_from_objective(obj_str: str) -> Optional[str]:
    """
    In old-style 'Channels / Objectives' templates, the objective cell contains
    values like "Video Views", "Reach", "Awareness" which are always Meta-family.
    Returns a platform string or None if we can't infer.
    """
    obj_lower = obj_str.lower()
    # These objectives in a "Channels/Objectives" column imply Meta (default)
    meta_objectives = {
        "video views", "reach", "awareness", "engagement", "engagements",
        "post engagement", "traffic", "clicks", "leads", "page like", "page likes",
        "recch",  # typo of "reach"
        "follower", "ig follower",
    }
    for kw in meta_objectives:
        if kw in obj_lower:
            return "Meta"
    return None


# ---------------------------------------------------------------------------
# Plan row normalizer
# ---------------------------------------------------------------------------
def normalize_row(raw: dict[str, Any], source_file: str, row_num: int,
                  issues: list[dict], new_platforms: set[str]) -> Optional[dict[str, Any]]:
    """Convert a raw row dict into a normalized row dict."""
    platform_raw = raw.get("platform")

    # "Channels / Objectives" templates: "objective" column doubles as platform indicator
    # The objective value IS the platform+objective (e.g., "Video Views" → meta_ig + awareness)
    if platform_raw is None:
        obj_raw_fallback = raw.get("objective")
        if obj_raw_fallback:
            # Infer platform from objective label for "Channels/Objectives" style sheets
            obj_str = str(obj_raw_fallback).strip().lower()
            platform_inferred = _infer_platform_from_objective(obj_str)
            if platform_inferred:
                platform_raw = platform_inferred
            else:
                return None
        else:
            return None

    platform_str = str(platform_raw).strip()
    if not platform_str or platform_str.lower() in ("–", "-", "—"):
        return None

    platform, is_new = normalize_platform(platform_str)
    if platform is None:
        # Value is an objective/KPI label — not a platform row
        return None
    if is_new:
        new_platforms.add(platform)
        log.debug(f"  New platform '{platform_str}' → '{platform}' in {source_file} row {row_num}")

    # Budget
    budget_raw = raw.get("budget") or raw.get("media_spend") or raw.get("total_budget")
    budget = parse_number(budget_raw)

    # Percentage
    pct_raw = raw.get("percentage")
    percentage = parse_pct(pct_raw) if pct_raw is not None else None

    # KPIs
    def kpi_range(key: str) -> tuple[Optional[float], Optional[float]]:
        val = raw.get(key)
        if val is None:
            return None, None
        lo, hi = parse_range(val)
        return lo, hi

    reach_lo, reach_hi = kpi_range("reach")
    impressions_lo, impressions_hi = kpi_range("impressions")
    clicks_lo, clicks_hi = kpi_range("clicks")
    leads_lo, leads_hi = kpi_range("leads")
    video_views_lo, video_views_hi = kpi_range("video_views")
    engagements_lo, engagements_hi = kpi_range("engagements")
    freq_lo, freq_hi = kpi_range("frequency")
    cpm_lo, cpm_hi = kpi_range("cpm")
    cpc_lo, cpc_hi = kpi_range("cpc")
    cpl_lo, cpl_hi = kpi_range("cpl")
    cpr_lo, cpr_hi = kpi_range("cpr")
    cpv_lo, cpv_hi = kpi_range("cpv")
    ctr_lo, ctr_hi = kpi_range("ctr")
    lpv_lo, lpv_hi = kpi_range("landing_page_views")
    page_likes_lo, page_likes_hi = kpi_range("page_likes")

    # Objective normalization — prefer explicit "objective" column over platform
    obj_raw = raw.get("objective")
    objective = normalize_objective(obj_raw)

    # Audience type inference
    audience_size_raw = raw.get("audience_size")
    audience_type = infer_audience_type(audience_size_raw)

    # Build initial kpis dict
    kpis: dict[str, Optional[float]] = {
        "reach_low": reach_lo, "reach_high": reach_hi,
        "impressions_low": impressions_lo, "impressions_high": impressions_hi,
        "clicks_low": clicks_lo, "clicks_high": clicks_hi,
        "leads_low": leads_lo, "leads_high": leads_hi,
        "video_views_low": video_views_lo, "video_views_high": video_views_hi,
        "engagements_low": engagements_lo, "engagements_high": engagements_hi,
        "frequency_low": freq_lo, "frequency_high": freq_hi,
        "cpm_low": cpm_lo, "cpm_high": cpm_hi,
        "cpc_low": cpc_lo, "cpc_high": cpc_hi,
        "cpl_low": cpl_lo, "cpl_high": cpl_hi,
        "cpr_low": cpr_lo, "cpr_high": cpr_hi,
        "cpv_low": cpv_lo, "cpv_high": cpv_hi,
        "ctr_low": ctr_lo, "ctr_high": ctr_hi,
        "landing_page_views_low": lpv_lo, "landing_page_views_high": lpv_hi,
        "page_likes_low": page_likes_lo, "page_likes_high": page_likes_hi,
    }

    # Enhancement 5: Apply avg_cost based on objective
    avg_cost_raw = raw.get("avg_cost")
    if avg_cost_raw is not None:
        ac_lo, ac_hi = parse_range(avg_cost_raw)
        if objective in (None, "awareness") and kpis.get("cpm_low") is None:
            kpis["cpm_low"] = ac_lo
            kpis["cpm_high"] = ac_hi
        elif objective == "traffic" and kpis.get("cpc_low") is None:
            kpis["cpc_low"] = ac_lo
            kpis["cpc_high"] = ac_hi
        elif objective == "engagement" and kpis.get("cpe_low") is None:
            kpis["cpe_low"] = ac_lo
            kpis["cpe_high"] = ac_hi
        elif objective in ("video_views", "awareness") and kpis.get("cpv_low") is None:
            # If platform suggests video and awareness, treat as CPV
            plat_lower = platform.lower()
            if "youtube" in plat_lower or "video" in plat_lower:
                kpis["cpv_low"] = ac_lo
                kpis["cpv_high"] = ac_hi
        elif objective == "leads" and kpis.get("cpl_low") is None:
            kpis["cpl_low"] = ac_lo
            kpis["cpl_high"] = ac_hi

    # Enhancement 2: Extract KPIs from estimated_deliveries text
    est_del = raw.get("estimated_deliveries")
    if est_del and isinstance(est_del, str):
        text_kpis = extract_kpis_from_text(est_del)
        for k, v in text_kpis.items():
            if v is not None and kpis.get(k) is None:
                kpis[k] = v

    # Enhancement 4: Second pass — scan ALL text cell values for embedded KPI data
    for cell_val in raw.values():
        if isinstance(cell_val, str) and len(cell_val) > 8:
            parsed = extract_kpis_from_text(cell_val)
            for k, v in parsed.items():
                if v is not None and kpis.get(k) is None:
                    kpis[k] = v

    return {
        "platform": platform,
        "platform_raw": platform_str,
        "objective": objective,
        "audience_type": audience_type,
        "audience_name": str(raw["audience_name"]).strip() if raw.get("audience_name") else None,
        "audience_size": str(audience_size_raw).strip() if audience_size_raw else None,
        "targeting": str(raw["targeting"]).strip()[:500] if raw.get("targeting") else None,
        "creative": str(raw["creative"]).strip()[:200] if raw.get("creative") else None,
        "country": str(raw["country"]).strip() if raw.get("country") else None,
        "buy_type": str(raw["buy_type"]).strip() if raw.get("buy_type") else None,
        "campaign_period": str(raw["campaign_period"]).strip() if raw.get("campaign_period") else None,
        "budget": budget,
        "percentage": percentage,
        "kpis": kpis,
    }


def normalize_objective(raw: Any) -> Optional[str]:
    if raw is None:
        return None
    s = str(raw).strip().lower()
    if any(kw in s for kw in ["awareness", "reach", "video view", "views", "cpm"]):
        return "awareness"
    if any(kw in s for kw in ["engagement", "engage", "page like", "follower", "like"]):
        return "engagement"
    if any(kw in s for kw in ["traffic", "click", "cpc", "website"]):
        return "traffic"
    if any(kw in s for kw in ["lead", "cpl", "conversion", "result"]):
        return "leads"
    return None


def infer_audience_type(audience_size: Any) -> Optional[str]:
    """Infer 'mass' or 'niche' from audience size string."""
    if audience_size is None:
        return None
    lo, hi = parse_range(audience_size)
    size = lo or hi
    if size is None:
        return None
    if size >= 1_000_000:
        return "mass"
    return "niche"


# ---------------------------------------------------------------------------
# Fee extraction
# ---------------------------------------------------------------------------
def extract_fees(ws, header_row: int) -> dict[str, Any]:
    """Look for management fee / ASP fee info near the data."""
    fees: dict[str, Any] = {
        "fee_1_pct": None, "fee_1_label": None,
        "fee_2_pct": None, "fee_2_label": None,
        "currency": "LKR",
        "total_budget": None,
        "media_spend": None,
    }
    for r_idx in range(1, min(header_row, 15)):
        row = tuple(cell.value for cell in ws[r_idx])
        for c_idx, cell in enumerate(row):
            if cell is None:
                continue
            s = str(cell).strip().lower()
            if "lkr" in s or "values are in lkr" in s:
                fees["currency"] = "LKR"
            if "usd" in s:
                fees["currency"] = "USD"
            if "management fee" in s or "mgmt fee" in s or "mgt fee" in s:
                # Try the next cell
                for offset in range(1, 4):
                    if c_idx + offset < len(row) and row[c_idx + offset] is not None:
                        pct = parse_pct(row[c_idx + offset])
                        if pct is not None:
                            fees["fee_1_pct"] = pct
                            fees["fee_1_label"] = "Management Fee"
                            break
            if "asp" in s and "fee" in s:
                for offset in range(1, 4):
                    if c_idx + offset < len(row) and row[c_idx + offset] is not None:
                        pct = parse_pct(row[c_idx + offset])
                        if pct is not None:
                            fees["fee_2_pct"] = pct
                            fees["fee_2_label"] = "ASP Fee"
                            break

    # Scan below data for fee annotations
    for r_idx in range(header_row, min(ws.max_row + 1, header_row + 20)):
        row = tuple(cell.value for cell in ws[r_idx])
        for c_idx, cell in enumerate(row):
            if cell is None:
                continue
            s = str(cell).strip().lower()
            if "management fee" in s or "mgmt fee" in s:
                for offset in range(1, 4):
                    if c_idx + offset < len(row) and row[c_idx + offset] is not None:
                        pct = parse_pct(row[c_idx + offset])
                        if pct is not None and fees["fee_1_pct"] is None:
                            fees["fee_1_pct"] = pct
                            fees["fee_1_label"] = "Management Fee"

    return fees


# ---------------------------------------------------------------------------
# Budget extraction from metadata cell (e.g. "300,000 (Including management fee 15%)")
# ---------------------------------------------------------------------------
_BUDGET_RE = re.compile(r"([\d,]+(?:\.\d+)?)", re.IGNORECASE)
_FEE_RE = re.compile(r"(\d+(?:\.\d+)?)\s*%", re.IGNORECASE)


def parse_budget_cell(val: Any) -> tuple[Optional[float], Optional[float]]:
    """
    Parse cells like '300,000\n(Including management fee 15%)'.
    Returns (total_budget, fee_pct).
    """
    if val is None:
        return None, None
    s = str(val).replace("\n", " ")
    budget_match = _BUDGET_RE.search(s)
    fee_match = _FEE_RE.search(s)
    budget = parse_number(budget_match.group(1)) if budget_match else None
    fee_pct = float(fee_match.group(1)) if fee_match else None
    return budget, fee_pct


# ---------------------------------------------------------------------------
# Process a single worksheet as a media plan
# ---------------------------------------------------------------------------
def process_media_plan_sheet(
    ws,
    source_file: str,
    client: str,
    product: str,
    issues: list[dict],
    new_platforms: set[str],
    new_creatives: set[str],
) -> Optional[dict[str, Any]]:
    header_row, col_map = find_header_row(ws)
    if header_row < 0 or len(col_map) < 2:
        log.warning(f"  → Could not detect header in sheet '{ws.title}'")
        return None

    log.info(f"  → Template detected: header at row {header_row}, {len(col_map)} cols: {list(col_map.values())}")

    # Extract metadata above header
    meta = extract_metadata(ws, header_row)
    fees = extract_fees(ws, header_row)

    # Try to get campaign name / budget from first column or metadata
    campaign_name = meta.get("campaign_name")
    reference_number = meta.get("reference_number")
    prepared_by = meta.get("prepared_by")
    date_str = meta.get("date")

    # Check if campaign_name column exists in data
    campaign_col = next((k for k, v in col_map.items() if v == "campaign_name"), None)
    total_budget_col = next((k for k, v in col_map.items() if v == "total_budget"), None)
    budget_col = next((k for k, v in col_map.items() if v in ("budget", "media_spend")), None)

    # Try to read campaign budget from first data row's total_budget or budget cell
    first_data_row = header_row + 1
    row_vals = tuple(cell.value for cell in ws[first_data_row]) if first_data_row <= ws.max_row else ()

    if total_budget_col and total_budget_col <= len(row_vals):
        tb, fee_pct = parse_budget_cell(row_vals[total_budget_col - 1])
        if tb:
            fees["total_budget"] = tb
        if fee_pct and fees["fee_1_pct"] is None:
            fees["fee_1_pct"] = fee_pct
            fees["fee_1_label"] = "Management Fee"
    elif budget_col and budget_col <= len(row_vals):
        tb, fee_pct = parse_budget_cell(row_vals[budget_col - 1])
        if tb:
            fees["total_budget"] = tb

    # Extract all data rows
    raw_rows = extract_rows(ws, header_row, col_map)
    if not raw_rows:
        log.warning(f"  → No data rows extracted from sheet '{ws.title}'")
        return None

    # Try to get campaign name from first data row if not in metadata
    if not campaign_name and raw_rows:
        first = raw_rows[0]
        if first.get("campaign_name"):
            campaign_name = str(first["campaign_name"]).strip()

    # Campaign period from first row
    campaign_period = None
    if raw_rows:
        campaign_period = raw_rows[0].get("campaign_period")
        if campaign_period:
            campaign_period = str(campaign_period).strip()

    # Normalize each row
    normalized_rows = []
    for i, row in enumerate(raw_rows):
        normed = normalize_row(row, source_file, header_row + 1 + i, issues, new_platforms)
        if normed:
            if normed.get("creative") and normed["creative"] not in ("–", "-"):
                new_creatives.add(normed["creative"])
            normalized_rows.append(normed)

    if not normalized_rows:
        return None

    # Determine currency from fees or column headers
    currency = fees.get("currency", "LKR")

    # Compute media spend from total budget + fee if possible
    total_budget = fees.get("total_budget")
    media_spend = fees.get("media_spend")
    fee_pct = fees.get("fee_1_pct")

    if total_budget and fee_pct and not media_spend:
        media_spend = round(total_budget / (1 + fee_pct / 100), 2)
    elif not total_budget:
        # Sum row budgets
        row_budgets = [r["budget"] for r in normalized_rows if r.get("budget")]
        if row_budgets:
            total_budget = sum(row_budgets)

    # Campaign period from most common row value
    if not campaign_period and normalized_rows:
        periods = [r.get("campaign_period") for r in normalized_rows if r.get("campaign_period")]
        if periods:
            campaign_period = max(set(periods), key=periods.count)

    return {
        "source_file": source_file,
        "source_sheet": ws.title,
        "client": client,
        "product": product,
        "campaign_name": campaign_name,
        "reference_number": reference_number,
        "prepared_by": prepared_by,
        "date": date_str,
        "campaign_period": campaign_period,
        "total_budget": total_budget,
        "media_spend": media_spend,
        "fee_1_pct": fees.get("fee_1_pct"),
        "fee_1_label": fees.get("fee_1_label"),
        "fee_2_pct": fees.get("fee_2_pct"),
        "fee_2_label": fees.get("fee_2_label"),
        "currency": currency,
        "rows": normalized_rows,
        "notes": None,
    }


# ---------------------------------------------------------------------------
# Process a KPI/performance sheet
# ---------------------------------------------------------------------------
def process_kpi_sheet(
    ws,
    source_file: str,
    client: str,
    issues: list[dict],
) -> Optional[dict[str, Any]]:
    header_row, col_map = find_header_row(ws)
    if header_row < 0 or len(col_map) < 2:
        return None

    meta = extract_metadata(ws, header_row)
    raw_rows = extract_rows(ws, header_row, col_map)
    if not raw_rows:
        return None

    campaign_name = meta.get("campaign_name")
    if not campaign_name and raw_rows:
        campaign_name = raw_rows[0].get("campaign_name")
        if campaign_name:
            campaign_name = str(campaign_name).strip()

    entries = []
    for row in raw_rows:
        platform_raw = row.get("platform")
        if not platform_raw:
            continue
        platform, _ = normalize_platform(str(platform_raw).strip())

        entry = {
            "platform": platform,
            "platform_raw": str(platform_raw).strip(),
            "actual_impressions": parse_number(row.get("impressions")),
            "actual_reach": parse_number(row.get("reach")),
            "actual_clicks": parse_number(row.get("clicks")),
            "actual_spend": parse_number(row.get("budget") or row.get("media_spend")),
            "actual_video_views": parse_number(row.get("video_views")),
            "actual_engagements": parse_number(row.get("engagements")),
            "actual_leads": parse_number(row.get("leads")),
            "actual_cpm": parse_number(row.get("cpm")),
            "actual_cpc": parse_number(row.get("cpc")),
            "actual_cpl": parse_number(row.get("cpl")),
            "actual_frequency": parse_number(row.get("frequency")),
        }
        if any(v is not None for k, v in entry.items() if k != "platform"):
            entries.append(entry)

    if not entries:
        return None

    return {
        "source_file": source_file,
        "source_sheet": ws.title,
        "client": client,
        "campaign_name": campaign_name,
        "period_label": meta.get("date"),
        "entries": entries,
    }


# ---------------------------------------------------------------------------
# File-level processing
# ---------------------------------------------------------------------------
def should_skip_file(path: Path) -> tuple[bool, str]:
    name = path.name
    if name.startswith("~$"):
        return True, "lock file"
    if path.suffix.lower() in (".pdf", ".pptx", ".ods", ".ppt", ".csv"):
        return True, f"unsupported type ({path.suffix})"
    if ":Zone.Identifier" in str(path):
        return True, "Zone.Identifier metadata file"
    return False, ""


_PERIOD_PATTERN = re.compile(
    r"^(january|february|march|april|may|june|july|august|september|october|november|december"
    r"|\d{4}|jan|feb|mar|apr|jun|jul|aug|sep|oct|nov|dec)",
    re.IGNORECASE,
)


# Keyword → product name map for Peoples Bank (ordered: more specific first)
_PB_PRODUCT_KEYWORDS: list[tuple[list[str], str]] = [
    (["vehicle loan", "vehicle loan leads", "vehicle loan video"], "Vehicle Loan"),
    (["housing loan", "housing loan google", "housing loan meta"], "Housing Loan"),
    (["professional loan", "professional-loan"], "Professional Loan"),
    (["sme loan", "sme day", "sme sector", "sme leads", "revitalisation & revival sme", "revitalisation revival sme"], "SME Loan"),
    (["business loan", "biz"], "Business Loan"),
    (["remittance", "vasi kotiyai", "november-remittance", "remittance campaign", "ethera hamuwa", "shrama harasara", "sanwathsara pranama"], "Remittance"),
    (["christmas credit card", "christmas campaign", "christmas"], "Christmas Campaign"),
    (["vanitha vasana credit card", "vanitha vasana", "vanitha saviya"], "Vanitha Vasana Credit Card"),
    (["card offer", "card offers", "avurudu", "havelock city mall credit card", "credit card", "0_ credit card", "0 instalment"], "Credit Cards"),
    (["leasing"], "Leasing"),
    (["peoples pay", "pay app", "pay yt", "pay ytp"], "Pay App"),
    (["page like campaign", "fb likes", "fb page likes", "page like", "fb like"], "Page Like Campaign"),
    (["anniversary", "64th anniversary"], "Anniversary Campaign"),
    (["women's day", "women day", "women_s day"], "Women's Day"),
    (["yes teen", "yesteen", "idb biz teen"], "Yes Teen"),
    (["4m digital", "4 m digital"], "Digital Customers Campaign"),
    (["best web"], "Web Campaign"),
    (["disaster relief", "disaster fund"], "CSR Campaign"),
    (["ubereats", "uber eats"], "UberEats Promotion"),
    (["birth freedom"], "Birth Freedom Campaign"),
    (["people_s tower", "peoples tower"], "Tower Opening"),
    (["doodaru paranama", "dooDaru"], "DooDaru Campaign"),
    (["children_s day", "children day"], "Children's Day Campaign"),
    (["ethera udana"], "Ethera Udana Campaign"),
    (["digital ticketing", "digital ticket"], "Digital Ticketing"),
    (["technnovation", "technovation"], "Technovation Campaign"),
    (["shangri-la", "shangri la"], "Shangri-La Promotion"),
    (["yt campaign", "yt video campaign", "youtube campaign"], "YouTube Campaign"),
    (["tv program", "tv programme", "ආදරණීය", "වින්දනීය", "ශනිදා", "හත්වෙනි"], "TV Program Boosting"),
    (["empowering tomorrow", "innovators"], "Innovators Campaign"),
    (["sme loan leads", "sme loan leads"], "SME Loan"),
    (["1st day in 1st lesson", "first lesson"], "CSR Campaign"),
    (["always on brand", "brand recall"], "Brand Awareness"),
    (["ceo interview", "contribution for sme"], "Corporate Campaign"),
]


def _infer_product_from_filename(filename: str, client: str) -> str | None:
    """Try to infer product name from filename using keyword matching."""
    name_lower = filename.lower()
    if client == "Peoples Bank":
        for keywords, product_name in _PB_PRODUCT_KEYWORDS:
            if any(kw in name_lower for kw in keywords):
                return product_name
    return None


def infer_client_product(path: Path) -> tuple[str, str]:
    """Infer client and product from the directory structure."""
    parts = path.parts
    raw_idx = next((i for i, p in enumerate(parts) if p == "raw"), -1)
    if raw_idx == -1:
        return "Unknown", "Unknown"

    remaining = parts[raw_idx + 1:]
    if len(remaining) == 0:
        return "Unknown", "Unknown"
    client = remaining[0]

    # Normalize common client name variations
    if "peoples" in client.lower() or "people" in client.lower():
        client = "Peoples Bank"
    elif "fonterra" in client.lower():
        client = "Fonterra"

    if len(remaining) >= 2:
        product_dir = remaining[1]
        pd_lower = product_dir.lower()
        # Classify the subdirectory
        if "kpi" in pd_lower:
            # KPI/performance folder — try keyword inference from filename
            product = _infer_product_from_filename(path.stem, client)
        elif _PERIOD_PATTERN.match(product_dir):
            # Date-based directory (e.g. "January 2026", "March 2026") — try keyword inference
            product = _infer_product_from_filename(path.stem, client)
        elif path.suffix.lower() in (".xlsx",) and len(remaining) == 2:
            # File is directly inside the client folder with no product subfolder
            product = _infer_product_from_filename(path.stem, client)
        else:
            product = product_dir
    else:
        # File directly in client folder — try keyword inference
        product = _infer_product_from_filename(path.stem, client)

    return client, product or "Unknown"


def process_file(
    path: Path,
    issues: list[dict],
    new_platforms: set[str],
    new_creatives: set[str],
) -> dict[str, list]:
    """Process a single Excel file. Returns {'plans': [...], 'actuals': [...]}."""
    result: dict[str, list] = {"plans": [], "actuals": []}
    rel_path = str(path.relative_to(RAW_DIR))
    client, product = infer_client_product(path)

    log.info(f"Processing: {rel_path}")

    try:
        wb = load_workbook(str(path), data_only=True, read_only=False)
    except Exception as e:
        log.error(f"  → Failed to open workbook: {e}")
        issues.append({"file": rel_path, "issue": f"Cannot open: {e}"})
        return result

    # Enhancement 3: If file lives inside a "KPI" folder, bias toward kpi_performance
    parent_lower = path.parent.name.lower()
    filename_lower = path.name.lower()
    is_kpi_context = "kpi" in parent_lower or "kpi" in filename_lower

    for sheet_name in wb.sheetnames:
        try:
            ws = wb[sheet_name]
            file_type = classify_file(ws)
            # Override unknown classification for files in KPI folders
            if file_type == "unknown" and is_kpi_context:
                file_type = "kpi_performance"
                log.debug(f"  Sheet '{sheet_name}' → kpi_performance (KPI folder override)")
            else:
                log.debug(f"  Sheet '{sheet_name}' → {file_type}")

            if file_type == "media_plan":
                plan = process_media_plan_sheet(
                    ws, rel_path, client, product, issues, new_platforms, new_creatives
                )
                if plan:
                    result["plans"].append(plan)
                    log.info(f"  ✓ Sheet '{sheet_name}': {len(plan['rows'])} rows extracted")

            elif file_type == "kpi_performance":
                actual = process_kpi_sheet(ws, rel_path, client, issues)
                if actual:
                    result["actuals"].append(actual)
                    log.info(f"  ✓ Sheet '{sheet_name}': {len(actual['entries'])} KPI entries")

            else:
                log.debug(f"  → Sheet '{sheet_name}': unknown type, skipping")

        except Exception as e:
            log.error(f"  → Sheet '{sheet_name}': unexpected error: {e}")
            issues.append({"file": rel_path, "sheet": sheet_name, "issue": str(e)})

    wb.close()
    return result


# ---------------------------------------------------------------------------
# New entity discovery
# ---------------------------------------------------------------------------
KNOWN_PRODUCTS = {
    # Fonterra
    "Anchor PediaPro", "Anchor Hot Chocolate", "Anchor Milk", "Anchor Newdale",
    # Ritzbury
    "Deckers", "Choco Bar",
    # Peoples Bank — original seed
    "Credit Cards", "Leasing", "Remittance", "FX-FUTURE", "Christmas Campaign", "SME",
    # Peoples Bank — additional products inferred from files
    "Vehicle Loan", "Housing Loan", "Professional Loan", "SME Loan", "Business Loan",
    "Christmas Campaign", "Vanitha Vasana Credit Card", "Pay App", "Page Like Campaign",
    "Anniversary Campaign", "Women's Day", "Yes Teen", "Digital Customers Campaign",
    "Web Campaign", "CSR Campaign", "UberEats Promotion", "Birth Freedom Campaign",
    "Tower Opening", "DooDaru Campaign", "Children's Day Campaign", "Ethera Udana Campaign",
    "Digital Ticketing", "Technovation Campaign", "Shangri-La Promotion", "YouTube Campaign",
    "TV Program Boosting", "Innovators Campaign", "Brand Awareness", "Corporate Campaign",
    "Card Offers", "Business Loan",
}


def discover_new_entities(
    plans: list[dict],
    new_platforms: set[str],
    new_creatives: set[str],
    client: str,
) -> dict[str, Any]:
    """Collect new platforms, creatives, and products not yet in PlanFlow."""
    products: list[dict] = []
    seen_products: set[str] = set()

    _SKIP_PRODUCTS = {"Unknown", "KPI Reports", ""}
    for plan in plans:
        p_client = plan.get("client", "")
        p_product = plan.get("product", "")
        key = f"{p_client}|{p_product}"
        if (p_product
                and p_product not in KNOWN_PRODUCTS
                and p_product not in _SKIP_PRODUCTS
                and not p_product.endswith(".xlsx")
                and not _PERIOD_PATTERN.match(p_product)
                and key not in seen_products):
            products.append({"client": p_client, "name": p_product})
            seen_products.add(key)

    return {
        "platforms": sorted(new_platforms),
        "creative_types": sorted(new_creatives),
        "products": products,
    }


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    global log
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    log = setup_logging()

    all_plans: list[dict] = []
    all_actuals: list[dict] = []
    issues: list[dict] = []
    new_platforms: set[str] = set()
    new_creatives: set[str] = set()

    stats = {
        "total_files_scanned": 0,
        "skipped_files": 0,
        "media_plans_found": 0,
        "kpi_sheets_found": 0,
        "unknown_files": 0,
        "errors": 0,
    }

    # Collect all files
    all_files = sorted(RAW_DIR.rglob("*"))
    xlsx_files = [
        f for f in all_files
        if f.is_file() and f.suffix.lower() == ".xlsx"
        and not f.name.startswith("~$")
        and ":Zone.Identifier" not in str(f)
    ]

    # Also log skipped files
    for f in all_files:
        if not f.is_file():
            continue
        skip, reason = should_skip_file(f)
        if skip:
            log.debug(f"Skipping: {f.relative_to(RAW_DIR)} ({reason})")
            stats["skipped_files"] += 1

    log.info(f"\n{'='*60}")
    log.info(f"PlanFlow Data Import Scanner")
    log.info(f"Raw directory: {RAW_DIR}")
    log.info(f"Files to process: {len(xlsx_files)}")
    log.info(f"{'='*60}\n")

    for path in xlsx_files:
        stats["total_files_scanned"] += 1
        rel = str(path.relative_to(RAW_DIR))

        skip, reason = should_skip_file(path)
        if skip:
            log.info(f"[SKIP] {rel}: {reason}")
            stats["skipped_files"] += 1
            continue

        try:
            file_result = process_file(path, issues, new_platforms, new_creatives)
            plans_found = len(file_result["plans"])
            actuals_found = len(file_result["actuals"])

            if plans_found == 0 and actuals_found == 0:
                stats["unknown_files"] += 1
                issues.append({"file": rel, "issue": "No extractable sheets found"})

            all_plans.extend(file_result["plans"])
            all_actuals.extend(file_result["actuals"])
            stats["media_plans_found"] += plans_found
            stats["kpi_sheets_found"] += actuals_found

        except Exception as e:
            log.error(f"FATAL processing {rel}: {e}")
            stats["errors"] += 1
            issues.append({"file": rel, "issue": f"Fatal error: {e}"})

    # Remove TikTok/LinkedIn from new_platforms if they're in PLATFORM_NORMALIZE
    # (they ARE known to the normalizer, just not to PlanFlow yet)
    truly_new = {p for p in new_platforms if p not in KNOWN_PLATFORMS}

    # Discover new entities
    new_entities = discover_new_entities(all_plans, truly_new, new_creatives, "")

    # Build scan report
    scan_report = {
        "total_files_scanned": stats["total_files_scanned"],
        "skipped_files": stats["skipped_files"],
        "media_plans_found": stats["media_plans_found"],
        "kpi_sheets_found": stats["kpi_sheets_found"],
        "unknown_files": stats["unknown_files"],
        "errors": stats["errors"],
        "new_platforms_found": new_entities["platforms"],
        "new_creative_types_found": new_entities["creative_types"][:20],  # top 20
        "data_quality_issues": issues,
    }

    # Write outputs
    def write_json(filename: str, data: Any):
        outpath = OUT_DIR / filename
        with open(outpath, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False, default=str)
        log.info(f"  → {outpath}")

    log.info(f"\n{'='*60}")
    log.info("Writing output files...")
    write_json("scan-report.json", scan_report)
    write_json("plans.json", all_plans)
    write_json("actuals.json", all_actuals)
    write_json("new-entities.json", new_entities)

    # Print summary
    print(f"\n{'='*60}")
    print(f"SCAN COMPLETE")
    print(f"{'='*60}")
    print(f"  Files scanned:        {stats['total_files_scanned']}")
    print(f"  Media plan sheets:    {stats['media_plans_found']}")
    print(f"  KPI/actuals sheets:   {stats['kpi_sheets_found']}")
    print(f"  Unknown files:        {stats['unknown_files']}")
    print(f"  Errors:               {stats['errors']}")
    if truly_new:
        print(f"  ⚠  New platforms:     {', '.join(sorted(truly_new))}")
    print(f"\nOutput written to: {OUT_DIR}/")
    print(f"  scan-report.json   plans.json   actuals.json   new-entities.json")
    print(f"  Detailed log: {LOG_FILE}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
